from openpyxl import load_workbook

wb = load_workbook(filename = '2017-01-01_MMU_ubill_consume_month.xlsx')
ws = wb.active
d = ws.cell(row=4, column=2, value=10)
print(d.value)